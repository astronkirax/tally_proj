"""Client Masters — the trial-balance / opening data a CA supplies once per year.

Why this exists: for a commission / pass-through business, the bank payments are almost
all customer-and-driver settlements (Sundry Creditors), NOT P&L expenses — so a P&L can't
be summed from raw bank rows. Every CA therefore keeps a trial balance. The masters ARE
that trial balance (income treatment, expense ledgers, capital, fixed assets, opening
balances). AuditWedge assembles them into Tally-format statements and auto-fills /
cross-checks the items the bank + invoices CAN supply (Bank Charges, Gateway Charges,
bank closing balance) — while the Summary is derived purely from the bank.

`example_masters()` is pre-filled from the FY22-23 auditor docs so the whole pipeline can
be verified end-to-end. In the app the CA fills the equivalent Excel template.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def D(x) -> Decimal:
    return Decimal(str(x))


def _num(v, default: str = "0") -> Decimal:
    if v is None or str(v).strip() == "":
        return Decimal(default)
    try:
        return Decimal(str(v).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


def _pdate(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        y, m, d = str(v).split("-")
        return date(int(y), int(m), int(d))
    except (ValueError, TypeError):
        return None


@dataclass
class Masters:
    client_name: str
    gstin: str
    pan: str
    period_from: date
    period_to: date
    gst_treatment: str = "both"  # both | commission | cab
    itr_form: str = "ITR-5"  # ITR-4 | ITR-5 | ITR-6 (entity type)
    # P&L income (CA treatment)
    gst_sales: Decimal = Decimal("0")          # GST Sales / Services (= cab turnover)
    commission_income: Decimal = Decimal("0")
    indirect_income: Decimal = Decimal("0")     # rewards / cashbacks
    # P&L expense ledgers (trial balance) — EXCLUDING Gateway Charges (comes from invoices)
    pnl_expenses: list[tuple[str, Decimal]] = field(default_factory=list)   # (ledger, amount)
    # Balance-sheet masters
    capital: list[tuple[str, Decimal]] = field(default_factory=list)        # (name, amount)
    fixed_assets: list[tuple[str, Decimal, Decimal]] = field(default_factory=list)  # (name, cost, dep%)
    opening: list[tuple[str, str, Decimal]] = field(default_factory=list)   # (ledger, asset|liability, amount)
    opening_pnl: Decimal = Decimal("0")
    # monthly cab turnover (feeds the Summary's GST turnover line)
    monthly_turnover: list[tuple[str, Decimal]] = field(default_factory=list)

    @property
    def gst_turnover(self) -> Decimal:
        return sum((amt for _, amt in self.monthly_turnover), Decimal("0"))


def example_masters() -> Masters:
    """FY22-23 Gateway2Konaseema, taken from the auditor's P&L / Balance Sheet / Summary."""
    return Masters(
        client_name="Gateway2Konaseema",
        gstin="37AAJCG0643B1ZZ",
        pan="AAJCG0643B",
        period_from=date(2022, 4, 1),
        period_to=date(2023, 3, 31),
        gst_treatment="both",
        gst_sales=D(75470),
        commission_income=D("55612.57"),
        indirect_income=D("1496.20"),
        pnl_expenses=[
            ("Audit Fee", D(5000)),
            ("Bank Charges", D("6400.34")),
            ("Conveyance Charges", D(17300)),
            ("Digital Marketing Keywords", D(10000)),
            ("Electricity Charges", D("4372.64")),
            ("Repairs & Maintenance", D(13330)),
            ("Software Expenses", D(2159)),
            ("Telephone Expenses", D("19658.32")),
        ],
        capital=[("Ananth Sai Ram Capital A/c", D(51000)), ("Vamsi Capital A/c", D(51000))],
        fixed_assets=[("Intangible Assets", D(56450), D(0))],
        opening=[
            ("Loans & Advances (Asset)", "asset", D("69465.64")),
            ("Cash-in-Hand", "asset", D("50243.33")),
            ("TDS Receivable", "asset", D(605)),
            ("Audit Fee Payable", "liability", D(5000)),
            ("Other Payables", "liability", D(80600)),
            ("GST Payable", "liability", D("-7102.10")),
        ],
        opening_pnl=D(2174),
        monthly_turnover=[
            ("Jul-2022", D(14890)), ("Sep-2022", D(14800)),
            ("Nov-2022", D(19890)), ("Mar-2023", D(25890)),
        ],
    )


# --------------------------------------------------------------------------- Excel I/O
_HDR = Font(bold=True, color="FFFFFF")
_FILL = PatternFill("solid", fgColor="1F4E78")


def _sheet(wb, title, headers, rows, widths):
    ws = wb.create_sheet(title) if wb.sheetnames != ["Sheet"] or title != "Settings" else wb.active
    if ws.title == "Sheet":
        ws.title = title
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = _HDR
        ws.cell(1, c).fill = _FILL
    for r in rows:
        ws.append(list(r))
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    return ws


def write_template(m: Masters | None = None) -> bytes:
    """Generate the Client Masters Excel (blank, or pre-filled from ``m``)."""
    if m is None:
        m = Masters("", "", "", date(2024, 4, 1), date(2025, 3, 31))
    wb = Workbook()
    _sheet(wb, "Settings", ["Field", "Value"], [
        ("Client Name", m.client_name), ("GSTIN", m.gstin), ("PAN", m.pan),
        ("Period From (YYYY-MM-DD)", m.period_from.isoformat()),
        ("Period To (YYYY-MM-DD)", m.period_to.isoformat()),
        ("GST Treatment (both/commission/cab)", m.gst_treatment),
        ("ITR Form (ITR-4/ITR-5/ITR-6)", m.itr_form),
        ("Opening P&L", float(m.opening_pnl)),
        ("Commission Income", float(m.commission_income)),
        ("GST Sales / Services", float(m.gst_sales)),
        ("Indirect Income", float(m.indirect_income)),
    ], [36, 28])
    _sheet(wb, "Capital", ["Account Name", "Amount"],
           [(n, float(a)) for n, a in m.capital] or [("", "")], [34, 16])
    _sheet(wb, "Fixed Assets", ["Asset", "Cost", "Dep Rate %"],
           [(n, float(c), float(d)) for n, c, d in m.fixed_assets] or [("", "", "")], [30, 16, 12])
    _sheet(wb, "Opening Balances", ["Ledger", "Group (asset/liability)", "Amount"],
           [(led, grp, float(a)) for led, grp, a in m.opening] or [("", "", "")], [30, 22, 16])
    _sheet(wb, "Monthly Turnover", ["Month", "Cab Turnover"],
           [(mth, float(a)) for mth, a in m.monthly_turnover] or [("", "")], [16, 16])
    _sheet(wb, "Expenses", ["Ledger", "Amount"],
           [(led, float(a)) for led, a in m.pnl_expenses] or [("", "")], [30, 16])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def example_template() -> bytes:
    """The masters template pre-filled with the FY22-23 Gateway example."""
    return write_template(example_masters())


def read_masters(source) -> Masters:
    """Read a filled Client Masters Excel back into a :class:`Masters`."""
    if isinstance(source, (bytes, bytearray)):
        source = BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)

    settings: dict[str, object] = {}
    if "Settings" in wb.sheetnames:
        for row in wb["Settings"].iter_rows(min_row=2, values_only=True):
            if row and row[0] not in (None, ""):
                settings[str(row[0]).strip().lower()] = row[1]

    def s(prefix: str, default=""):
        for k, v in settings.items():
            if k.startswith(prefix.lower()):
                return v if v is not None else default
        return default

    def body(name: str):
        if name not in wb.sheetnames:
            return []
        return [r for r in wb[name].iter_rows(min_row=2, values_only=True) if r and r[0] not in (None, "")]

    return Masters(
        client_name=str(s("client name")),
        gstin=str(s("gstin")),
        pan=str(s("pan")),
        period_from=_pdate(s("period from")) or date(2024, 4, 1),
        period_to=_pdate(s("period to")) or date(2025, 3, 31),
        gst_treatment=str(s("gst treatment", "both")).strip() or "both",
        itr_form=str(s("itr form", "ITR-5")).strip() or "ITR-5",
        gst_sales=_num(s("gst sales")),
        commission_income=_num(s("commission income")),
        indirect_income=_num(s("indirect income")),
        opening_pnl=_num(s("opening p&l")),
        pnl_expenses=[(str(r[0]), _num(r[1])) for r in body("Expenses")],
        capital=[(str(r[0]), _num(r[1])) for r in body("Capital")],
        fixed_assets=[(str(r[0]), _num(r[1]), _num(r[2] if len(r) > 2 else 0)) for r in body("Fixed Assets")],
        opening=[(str(r[0]), str(r[1]).strip().lower(), _num(r[2])) for r in body("Opening Balances")],
        monthly_turnover=[(str(r[0]), _num(r[1])) for r in body("Monthly Turnover")],
    )
