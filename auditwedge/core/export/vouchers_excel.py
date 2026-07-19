"""Convert a Tally voucher XML (the import file) into a readable Excel voucher register.

Same content as the .xml Tally import, but human-readable for review/sign-off before
import: one row per voucher with its Dr ledger, Cr ledger, amount and narration.
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONEY_FMT = "#,##0.00"
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _fmt_date(yyyymmdd: str | None) -> str:
    if yyyymmdd and len(yyyymmdd) == 8 and yyyymmdd.isdigit():
        return f"{yyyymmdd[6:8]}-{yyyymmdd[4:6]}-{yyyymmdd[0:4]}"
    return yyyymmdd or ""


def tally_xml_to_excel(xml_bytes: bytes) -> bytes:
    """Parse Tally import XML -> Excel voucher register (bytes)."""
    root = ET.fromstring(xml_bytes)
    wb = Workbook()
    ws = wb.active
    ws.title = "Tally Vouchers"

    headers = ["#", "Voucher Type", "Date", "Dr Ledger", "Cr Ledger", "Amount", "Narration"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    row = 2
    for i, v in enumerate(root.iter("VOUCHER"), start=1):
        vtype = v.get("VCHTYPE") or (v.findtext("VOUCHERTYPENAME") or "")
        date = _fmt_date(v.findtext("DATE"))
        narration = v.findtext("NARRATION") or ""
        dr = cr = ""
        amount = 0.0
        for le in v.findall("ALLLEDGERENTRIES.LIST"):
            name = le.findtext("LEDGERNAME") or ""
            deemed = (le.findtext("ISDEEMEDPOSITIVE") or "").strip().lower()
            amt = le.findtext("AMOUNT") or "0"
            if deemed == "yes":       # the debit leg (negative amount in Tally)
                dr = name
            else:                      # the credit leg (positive amount)
                cr = name
                try:
                    amount = abs(float(amt.replace(",", "")))
                except ValueError:
                    amount = 0.0
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=vtype)
        ws.cell(row=row, column=3, value=date)
        ws.cell(row=row, column=4, value=dr)
        ws.cell(row=row, column=5, value=cr)
        a = ws.cell(row=row, column=6, value=amount)
        a.number_format = MONEY_FMT
        ws.cell(row=row, column=7, value=narration)
        row += 1

    for i, w in enumerate([5, 14, 12, 26, 26, 14, 60], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


if __name__ == "__main__":  # python -m core.export.vouchers_excel in.xml out.xlsx
    import sys

    with open(sys.argv[1], "rb") as f:
        data = tally_xml_to_excel(f.read())
    with open(sys.argv[2], "wb") as f:
        f.write(data)
    print("wrote", sys.argv[2], len(data), "bytes")
