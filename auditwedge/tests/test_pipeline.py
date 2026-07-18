"""End-to-end test of the analyze() pipeline + exports, using the byte-upload path
the Streamlit UI takes.
"""
import xml.etree.ElementTree as ET
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from core.export.tally_xml import build_tally_xml
from core.export.workpaper import build_workpaper
from core.pipeline import analyze

FIX = Path(__file__).parent / "fixtures"

# Fixtures are real bank data, kept OUT of the public repo. Skip if absent.
pytestmark = pytest.mark.skipif(
    not (FIX / "hdfc_sample.pdf").exists(), reason="real-data fixture not present"
)


@pytest.fixture(scope="module")
def result():
    pdf_bytes = (FIX / "hdfc_sample.pdf").read_bytes()
    xlsx_bytes = (FIX / "invoices_sample.xlsx").read_bytes()
    return analyze(pdf_bytes, xlsx_bytes, use_llm=False)


def test_pipeline_reconciles(result):
    assert result.recon.balance_chains is True
    assert result.recon.matches_statement_summary is True


def test_classification_covers_most_rows(result):
    assert result.stats["rate"] >= 0.75  # rules alone classify the bulk


def test_expected_red_flags_present(result):
    codes = {f.code for f in result.flags}
    assert "NEAR_ZERO_BALANCE" in codes  # the near-zero-balance HIGH flag
    assert "ADVANCE_UNVOUCHED" in codes  # unvouched driver advances
    assert "INVOICE_WRONG_PERIOD" in codes  # wrong-FY invoices vs statement period
    assert any(f.severity.value == "high" for f in result.flags)


def test_workpaper_export_is_valid(result):
    data = build_workpaper(result.statement, result.recon, result.flags, result.stats)
    wb = openpyxl.load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Cover", "Reconciliation", "Transactions", "Ledger Summary", "Exceptions"]
    assert wb["Transactions"].max_row == len(result.statement.transactions) + 1


def test_tally_xml_export_is_well_formed(result):
    data = build_tally_xml(result.statement)
    root = ET.fromstring(data)
    vouchers = root.findall(".//VOUCHER")
    assert len(vouchers) == len(result.statement.transactions)
    kinds = {v.get("VCHTYPE") for v in vouchers}
    assert kinds == {"Payment", "Receipt"}
